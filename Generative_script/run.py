"""
disc_diffusion/run.py
======================
Entry point for the discrete diffusion pipeline.

Ranking pipeline (when drift model provided):
    1. Generate N candidates via diffusion
    2. Predict drift (GATv2) for ALL candidates
    3. Rank by: 0.4*x_drift + 0.35*y_drift + 0.15*symmetry + 0.10*constructability
    4. Show drift per floor for all candidates
"""

from __future__ import annotations

import argparse
import logging
import math
from pathlib import Path

import torch

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(name)s  %(message)s",
)
logger = logging.getLogger("disc_diffusion.run")


# -----------------------------------------------------------------------
# Commands
# -----------------------------------------------------------------------

def cmd_build_masks(args) -> None:
    from disc_diffusion.data.mask_builder import MaskBuilder
    logger.info("Building masks from: %s", args.data_root)
    mb = MaskBuilder(data_root=args.data_root, on_threshold=args.on_thresh, off_threshold=args.off_thresh)
    mb.build_all(split_ratio=1.0)
    mb.save(args.out)
    logger.info("Masks saved to: %s", args.out)
    for key, m in mb.masks.items():
        print(f"  {key}: n={m.n_samples}  must_on={int(m.must_on.sum())}  must_off={int(m.must_off.sum())}  variable={int(m.variable.sum())}")


def cmd_train(args) -> None:
    from disc_diffusion.data.slot_dataset import create_dataloaders
    from disc_diffusion.model.denoiser   import WallSlotDenoiser
    from disc_diffusion.model.diffusion  import DiscreteDiffusion
    from disc_diffusion.model.surrogate  import StructuralSurrogate
    from disc_diffusion.training.trainer import DiffusionTrainer

    logger.info("Loading data from: %s", args.data_root)
    train_loader, val_loader, _ = create_dataloaders(
        data_root=args.data_root, masks_path=args.masks,
        batch_size=args.batch_size, num_workers=args.workers, max_floors=20,
    )
    denoiser  = WallSlotDenoiser(in_channels=6, base_ch=args.base_ch, cond_in_dim=5, cond_embed_dim=128, t_embed_dim=128)
    surrogate = StructuralSurrogate(cond_dim=5, base_ch=32)
    diffusion = DiscreteDiffusion(T=args.T, schedule=args.schedule, device="cuda" if torch.cuda.is_available() else "cpu")
    cfg = {"lr": args.lr, "epochs": args.epochs, "save_freq": 5, "patience": 15, "ckpt_dir": args.ckpt_dir, "train_surrogate": True}
    trainer = DiffusionTrainer(denoiser=denoiser, diffusion=diffusion, train_loader=train_loader, val_loader=val_loader, config=cfg, surrogate=surrogate)
    if args.resume:
        trainer.load(args.resume)
    trainer.train(args.epochs)


def cmd_generate(args) -> None:
    import time, csv
    _measure = getattr(args, "measure_timing", False)
    _t_startup_begin = time.perf_counter() if _measure else 0.0

    from disc_diffusion.data.mask_builder   import MaskBuilder
    from disc_diffusion.model.denoiser      import WallSlotDenoiser
    from disc_diffusion.model.diffusion     import DiscreteDiffusion
    from disc_diffusion.model.surrogate     import StructuralSurrogate
    from disc_diffusion.generation.pipeline import GenerationPipeline
    from datetime import datetime
    import json

    # Timing accumulators (seconds). Populated only when --measure_timing.
    _tm = {"startup": 0.0, "generation": 0.0, "predict": 0.0,
           "rerank": 0.0, "save": 0.0}

    device = "cuda" if torch.cuda.is_available() else "cpu"
    mask_builder = MaskBuilder.load(args.masks)
    denoiser  = WallSlotDenoiser(in_channels=6)
    surrogate = StructuralSurrogate()
    diffusion = DiscreteDiffusion(T=200, device=device)
    ckpt = torch.load(args.ckpt, map_location="cpu")
    denoiser.load_state_dict(ckpt["denoiser"])
    if "surrogate" in ckpt:
        surrogate.load_state_dict(ckpt["surrogate"])

    # Generate candidates from diffusion
    pipeline = GenerationPipeline(denoiser, surrogate, diffusion, mask_builder, device=device)

    if _measure:
        if device == "cuda":
            torch.cuda.synchronize()
        _tm["startup"] = time.perf_counter() - _t_startup_begin
        _t0 = time.perf_counter()

    result = pipeline.generate(
        lx=args.lx, ly=args.ly, num_floors=args.floors,
        shear_ratio=args.shear if args.shear > 0 else None,
        num_candidates=args.candidates,
    )

    if _measure:
        if device == "cuda":
            torch.cuda.synchronize()
        _tm["generation"] = time.perf_counter() - _t0

    all_candidates = result["all"]

    # ---- Create output folder EARLY (needed for per-candidate workspace) ----
    timestamp   = datetime.now().strftime("%Y%m%d_%H%M%S")
    shear_tag   = f"s{int(args.shear * 100):03d}" if args.shear > 0 else "sauto"
    folder_name = f"plan_{args.lx}x{args.ly}_f{args.floors}_{shear_tag}_{timestamp}"
    out_dir     = Path(args.output_dir) / folder_name
    out_dir.mkdir(parents=True, exist_ok=True)

    # ---- Strict canonical drift prediction & re-ranking ----
    use_drift = False
    if (args.drift_checkpoint and args.graph_converter and args.inference_script
            and Path(args.drift_checkpoint).exists()
            and Path(args.graph_converter).exists()
            and Path(args.inference_script).exists()):
        per_cand_dir = out_dir / "per_candidate"
        per_cand_dir.mkdir(parents=True, exist_ok=True)

        all_candidates = _rerank_canonical_strict(
            all_candidates,
            lx=args.lx, ly=args.ly,
            num_floors=args.floors,
            shear_ratio=result["shear_ratio"],
            drift_ckpt=args.drift_checkpoint,
            graph_converter_path=args.graph_converter,
            inference_script_path=args.inference_script,
            per_cand_dir=per_cand_dir,
            timing_dict=(_tm if _measure else None),
            device=device,
        )
        use_drift = True
    else:
        # Fallback: use internal surrogate values (no canonical drift available)
        for cand in all_candidates:
            cand["drift_pred_norm"] = None
            cand["x_drift_mean"] = cand.get("x_dir_pred", 0)
            cand["y_drift_mean"] = cand.get("y_dir_pred", 0)
        logger.warning(
            "Canonical drift workflow not enabled. To enable, provide all "
            "three: --drift_checkpoint, --graph_converter, --inference_script. "
            "Using internal surrogate values for ranking (less accurate)."
        )

    # ---- Print final ranking ----
    best = all_candidates[0]
    print(result["report"])
    print(f"\nFinal ranking ({'canonical drift' if use_drift else 'internal surrogate'}):")
    for i, c in enumerate(all_candidates):
        print(f"  Rank #{i+1}: X_norm={c['x_drift_mean']:.2f}  Y_norm={c['y_drift_mean']:.2f}  "
              f"sym={c.get('symmetry_pred',0):.3f}  score={c.get('final_score',0):.4f}")

    if _measure:
        _t0 = time.perf_counter()

    # ---- Save result.json (best candidate, training-data format) ----
    best_drift_norm = best.get("drift_pred_norm", None)
    best_walls      = all_candidates[0]["walls"]
    floor_height    = 4.0

    story_details = []
    for floor_idx in range(args.floors):
        elevation = (floor_idx + 1) * floor_height
        z_bot     = elevation - floor_height
        z_top     = elevation

        if best_drift_norm is not None:
            norm_x_val = float(best_drift_norm[floor_idx][0])
            norm_y_val = float(best_drift_norm[floor_idx][1])
        else:
            norm_x_val = 0.0
            norm_y_val = 0.0

        wall_details = []
        for w_idx, w in enumerate(best_walls):
            cx, cy = w["cx"], w["cy"]
            ang    = w["angle"]
            half   = w.get("length", 2.0) / 2.0
            dx     = round(half * math.cos(math.radians(ang)), 3)
            dy     = round(half * math.sin(math.radians(ang)), 3)
            x1, y1 = round(cx - dx, 3), round(cy - dy, 3)
            x2, y2 = round(cx + dx, 3), round(cy + dy, 3)

            wall_details.append({
                "Element Name": w_idx + 1,
                "Object Type":  "Wall",
                "Object Label": f"W{w_idx + 1}",
                "Object Name":  w_idx + 1,
                "Elm Jt1": str([int(x1) if x1 == int(x1) else x1,
                          int(y1) if y1 == int(y1) else y1,
                          int(z_bot) if z_bot == int(z_bot) else z_bot]),
                "Elm Jt2": str([int(x2) if x2 == int(x2) else x2,
                          int(y2) if y2 == int(y2) else y2,
                          int(z_bot) if z_bot == int(z_bot) else z_bot]),
                "Elm Jt3": str([int(x2) if x2 == int(x2) else x2,
                          int(y2) if y2 == int(y2) else y2,
                          int(z_top) if z_top == int(z_top) else z_top]),
                "Elm Jt4": str([int(x1) if x1 == int(x1) else x1,
                          int(y1) if y1 == int(y1) else y1,
                          int(z_top) if z_top == int(z_top) else z_top]),
            })

        story_details.append({
            "Story":        f"Story{floor_idx + 1}",
            "Elevation":    round(elevation, 3),
            "Location":     "Top",
            "X-Dir":        norm_x_val,
            "Y-Dir":        norm_y_val,
            "Wall_Details": wall_details,
        })

    story_details = list(reversed(story_details))

    out_data = {
        "ProjectName": f"Project_{args.lx}x{args.ly}_{args.floors}",
        "File_details": [{
            "Name of file":     f"generated_{args.lx}x{args.ly}_f{args.floors}.json",
            "Shear_wall_ratio": float(result["shear_ratio"]),
            "Story_detail":     story_details,
        }],
    }
    with open(out_dir / "result.json", "w") as f:
        json.dump(out_data, f, indent=2)
    with open(out_dir / "report.txt", "w") as f:
        f.write(result["report"])

    # ---- Optional exports for the BEST candidate ----
    if args.save_before_json:
        json_path = out_dir / f"before_{args.lx}x{args.ly}_f{args.floors}.json"
        _export_before_json(
            walls=all_candidates[0]["walls"],
            lx=args.lx, ly=args.ly,
            num_floors=args.floors,
            shear_ratio=result["shear_ratio"],
            out_path=json_path,
        )
        print(f"  - before-format JSON: {json_path.name}")

        if args.save_etabs_xlsx:
            xlsx_path = out_dir / f"etabs_{args.lx}x{args.ly}_f{args.floors}_{shear_tag}.xlsx"
            _export_etabs_xlsx(json_path, xlsx_path,
                               template_path=getattr(args, "etabs_template", None))
            print(f"  - ETABS xlsx: {xlsx_path.name}")

    # ---- Save PNG per candidate ----
    for rank, cand in enumerate(all_candidates):
        _save_candidate_png(
            cand=cand, rank=rank + 1, is_best=(rank == 0),
            lx=args.lx, ly=args.ly, num_floors=args.floors,
            shear_ratio=result["shear_ratio"],
            out_dir=out_dir,
        )

    _save_comparison_table(all_candidates, args.lx, args.ly, out_dir, use_drift=use_drift)

    print(f"\nSaved to: {out_dir}")
    print(f"  - {len(all_candidates)} candidate PNGs")
    print(f"  - 1 comparison table")
    if use_drift:
        print(f"  - per_candidate/ (JSONs + graphs + canonical inference outputs)")

    # ---- Timing: close 'save' stage, append CSV row ----
    if _measure:
        _tm["save"] = time.perf_counter() - _t0
        # total_processing EXCLUDES save (PNG rendering is heavy; pending
        # supervisor decision). t_save still recorded separately, and
        # t_total_with_save kept for reference.
        total_proc = (_tm["generation"] + _tm["predict"]
                      + _tm["rerank"])
        total_with_save = total_proc + _tm["save"]
        timing_dir = getattr(args, "timing_dir", None) or args.output_dir
        Path(timing_dir).mkdir(parents=True, exist_ok=True)
        csv_path = Path(timing_dir) / "timing_log.csv"
        header = ["timestamp", "plan", "lx", "ly", "floors", "candidates",
                  "shear_target", "shear_achieved",
                  "t_startup", "t_generation", "t_predict", "t_rerank",
                  "t_save", "t_total_processing", "t_total_with_save"]
        row = [
            datetime.now().strftime("%Y-%m-%d_%H:%M:%S"),
            f"{args.lx}x{args.ly}", args.lx, args.ly, args.floors,
            args.candidates,
            f"{args.shear:.3f}",
            f"{float(result['shear_ratio']):.3f}",
            f"{_tm['startup']:.4f}", f"{_tm['generation']:.4f}",
            f"{_tm['predict']:.4f}", f"{_tm['rerank']:.4f}",
            f"{_tm['save']:.4f}", f"{total_proc:.4f}", f"{total_with_save:.4f}",
        ]
        write_header = not csv_path.exists()
        with open(csv_path, "a", newline="") as f:
            w = csv.writer(f)
            if write_header:
                w.writerow(header)
            w.writerow(row)
        print("\n" + "=" * 50)
        print("TIMING (seconds)")
        print(f"  startup            : {_tm['startup']:.3f}  (one-time, NOT in total)")
        print(f"  generation         : {_tm['generation']:.3f}")
        print(f"  predict            : {_tm['predict']:.3f}")
        print(f"  rerank             : {_tm['rerank']:.3f}")
        print(f"  save (PNG, etc.)   : {_tm['save']:.3f}  (NOT in total; pending review)")
        print(f"  TOTAL (gen+predict+rerank) : {total_proc:.3f}")
        print(f"  total incl. save           : {total_with_save:.3f}")
        print(f"  -> appended to {csv_path}")
        print("=" * 50)


# -----------------------------------------------------------------------
# Export best plan to ETABS-importable xlsx
# -----------------------------------------------------------------------

def _export_etabs_xlsx(json_path: Path, xlsx_path: Path) -> None:
    """
    Convert before-format JSON to ETABS-importable xlsx.
    Requires openpyxl: pip install openpyxl
    """
    try:
        import ast
        import uuid
        from openpyxl import Workbook
        from openpyxl.styles import Font

        with open(json_path) as f:
            data = json.load(f)

        fd      = data["File_details"][0]
        stories = sorted(fd["Story_detail"], key=lambda s: s["Elevation"])
        floor_h = 4.0

        # Infer layout
        max_x = max_y = 0
        for w in stories[0].get("Wall_Details", []):
            j1 = ast.literal_eval(str(w.get("Elm Jt1", "[0,0,0]")))
            j2 = ast.literal_eval(str(w.get("Elm Jt2", "[0,0,0]")))
            max_x = max(max_x, j1[0], j2[0])
            max_y = max(max_y, j1[1], j2[1])
        import math
        lx = max(1, int(math.ceil(max_x / 6)))
        ly = max(1, int(math.ceil(max_y / 6)))

        # Build joint registry
        joints = {}
        jc     = [1]

        def get_jid(coord):
            key = (round(coord[0],3), round(coord[1],3), round(coord[2],3))
            if key not in joints:
                joints[key] = jc[0]; jc[0] += 1
            return joints[key]

        wall_list = []
        story_elevations = []
        for story in stories:
            sname = story["Story"]; elev = story["Elevation"]
            story_elevations.append((sname, elev))
            for w in story.get("Wall_Details", []):
                j1 = ast.literal_eval(str(w["Elm Jt1"]))
                j2 = ast.literal_eval(str(w["Elm Jt2"]))
                j3 = ast.literal_eval(str(w["Elm Jt3"]))
                j4 = ast.literal_eval(str(w["Elm Jt4"]))
                wall_list.append({
                    "story": sname, "elem_name": w["Element Name"],
                    "obj_label": w["Object Label"], "obj_name": w["Object Name"],
                    "jt1": get_jid(j1), "jt2": get_jid(j2),
                    "jt3": get_jid(j3), "jt4": get_jid(j4),
                })

        id2coord = {v: k for k, v in joints.items()}

        def story_for_z(z):
            for sn, el in story_elevations:
                if abs(z - el) < 0.01: return sn
            return "Base"

        wb = Workbook()
        del wb["Sheet"]

        def mk(name, title, hdrs, units, rows):
            ws = wb.create_sheet(name)
            ws.append([f"TABLE:  {title}"])
            ws.append(hdrs); ws.append(units)
            for r in rows: ws.append(r)
            ws["A1"].font = Font(bold=True)

        # Grid
        xl = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
        grid_rows  = [[" G1","X (Cartesian)", xl[i], i*6, "End", "Yes"] for i in range(lx+1)]
        grid_rows += [["G1","Y (Cartesian)", str(j+1), j*6, "Start", "Yes"] for j in range(ly+1)]
        mk("Grid Definitions - Grid Lines","Grid Definitions - Grid Lines",
           ["Name","Grid Line Type","ID","Ordinate","Bubble Location","Visible"],
           [None,None,None,"m",None,None], grid_rows)

        # Stories (top first)
        colors = ["Yellow","Gray8Dark","Red","Blue","Green","Orange"]
        story_rows = []
        rev = list(reversed(story_elevations))
        master = rev[0][0]
        for i,(sn,_) in enumerate(rev):
            story_rows.append([sn, floor_h, "Yes" if i==0 else "No",
                               master if i>0 else None, "No",
                               colors[i%len(colors)], str(uuid.uuid4())])
        mk("Story Definitions","Story Definitions",
           ["Name","Height","Master Story","Similar To","Splice Story","Color","GUID"],
           [None,"m",None,None,None,None,None], story_rows)

        # Points
        pt_rows = [[jid,"No",story_for_z(id2coord[jid][2]),jid,"No",
                    int(id2coord[jid][0]) if id2coord[jid][0]==int(id2coord[jid][0]) else id2coord[jid][0],
                    int(id2coord[jid][1]) if id2coord[jid][1]==int(id2coord[jid][1]) else id2coord[jid][1],
                    int(id2coord[jid][2]) if id2coord[jid][2]==int(id2coord[jid][2]) else id2coord[jid][2],
                    None] for jid in range(1, len(joints)+1)]
        mk("Point Object Connectivity","Point Object Connectivity",
           ["UniqueName","Is Auto Point","Story","PointBay","IsSpecial","X","Y","Z","GUID"],
           [None,None,None,None,None,"m","m","m",None], pt_rows)

        # Restraints
        rc = 1
        rest_rows = []
        for jid in range(1, len(joints)+1):
            if id2coord[jid][2] < 0.01:
                rest_rows.append(["Base",jid,rc,"Yes","Yes","Yes","Yes","Yes","Yes"]); rc+=1
        mk("Joint Assigns - Restraints","Joint Assignments - Restraints",
           ["Story","Label","UniqueName","UX","UY","UZ","RX","RY","RZ"],
           [None]*9, rest_rows)

        # Wall Object Connectivity
        mk("Wall Object Connectivity","Wall Object Connectivity",
           ["UniqueName","Story","WallBay","UniquePt1","UniquePt2","UniquePt3","UniquePt4","Perimeter","Area","GUID"],
           [None]*10,
           [[w["elem_name"],w["story"],"Wall",w["jt1"],w["jt2"],w["jt3"],w["jt4"],None,None,None]
            for w in wall_list])

        # Objects and Elements - Areas
        mk("Objects and Elements - Areas","Objects and Elements - Areas",
           ["Story","Element Name","Object Type","Object Label","Object Name","Elm Jt1","Elm Jt2","Elm Jt3","Elm Jt4"],
           [None]*9,
           [[w["story"],w["elem_name"],"Wall",w["obj_label"],w["obj_name"],
             w["jt1"],w["jt2"],w["jt3"],w["jt4"]] for w in wall_list])

        # Area Assigns - Sect Prop
        mk("Area Assigns - Sect Prop","Area Assignments - Section Properties",
           ["Story","Label","UniqueName","Section Property","Property Type"],
           [None]*5,
           [[w["story"],"Wall",w["elem_name"],"W300","Wall"] for w in wall_list])

        # Fixed sheets
        mk("Program Control","Program Control",
           ["ProgramName","Version","ProgLevel","LicenseNum","CurrUnits",
            "StlFrmCode","CompBmCode","CompColCode","StlJstCode","ConcFrmCode","ConcSlbCode","ShrWallCode"],
           [None]*12,
           [["ETABS","21.2.0","Ultimate C","3010-*1BTXCCSZTGUALPH","kip, in, F",
             "AISC 360-16","AISC 360-16","AISC 360-16","SJI-2010","ACI 318-19","ACI 318-19","ACI 318-19"]])

        mk("Tower and Base Story Definition","Tower and Base Story Definitions",
           ["Tower","Color","GUID","Notes","BSName","BSElev","BSColor"],[None]*7,
           [["T1","Gray8Dark",str(uuid.uuid4()),None,"Base",0,"Gray8Dark"]])

        mk("Load Pattern Definitions","Load Pattern Definitions",
           ["Name","Is Auto Load","Type","Self Weight Multiplier","Auto Load","GUID"],[None]*6,
           [["~ChineseX","Yes","Other",0,None,str(uuid.uuid4())],
            ["~ChineseY","Yes","Other",0,None,str(uuid.uuid4())],
            ["~LLRF","Yes","Other",0,None,str(uuid.uuid4())],
            ["Dead","No","Dead",1,None,str(uuid.uuid4())],
            ["Live","No","Live",0,None,str(uuid.uuid4())],
            ["Super Dead Load","No","Super Dead",0,None,"1ac47084-22cf-4ef7-9577-8308fe7bedcc"],
            ["Wind","No","Wind",0,None,"65671afc-9585-4847-841d-bf6144d0a108"]])

        mk("Load Cases - Summary","Load Case Definitions - Summary",
           ["Name","Type","GUID"],[None]*3,
           [["Dead","Linear Static",str(uuid.uuid4())],
            ["1.2D+L+W","Linear Static","abdb0c4b-a9c4-4dc1-b422-76ab982aa8da"],
            ["Super Dead Load","Linear Static","797f76d0-e4d9-4e5a-b89c-0525aca1e2f7"],
            ["Wind","Linear Static","e962596b-8d4c-4da9-8dc2-47d352ebb8b6"],
            ["Live","Linear Static",str(uuid.uuid4())],
            ["Modal","Modal - Eigen",str(uuid.uuid4())]])

        mk("Load Cases - Linear Static","Load Case Definitions - Linear Static",
           ["Name","Exclude Group","Mass Source","Stiffness Type","Load Type",
            "Load Name","Load SF","Design Type","GUID","Notes"],[None]*10,
           [["1.2D+L+W",None,"MsSrc1","P-Delta","Load","Dead",1.2,"Program Determined",
             "abdb0c4b-a9c4-4dc1-b422-76ab982aa8da",None],
            ["1.2D+L+W",None,None,None,"Load","Super Dead Load",1.2,None,None,None],
            ["1.2D+L+W",None,None,None,"Load","Live",1,None,None,None],
            ["1.2D+L+W",None,None,None,"Load","Wind",1,None,None,None],
            ["Dead",None,"MsSrc1","P-Delta","Load","Dead",1,"Program Determined",str(uuid.uuid4()),None],
            ["Live",None,"MsSrc1","P-Delta","Load","Live",1,"Program Determined",str(uuid.uuid4()),None],
            ["Super Dead Load",None,"MsSrc1","P-Delta","Load","Super Dead Load",1,"Program Determined",
             "797f76d0-e4d9-4e5a-b89c-0525aca1e2f7",None],
            ["Wind",None,"MsSrc1","P-Delta","Load","Wind",1,"Program Determined",
             "e962596b-8d4c-4da9-8dc2-47d352ebb8b6",None]])

        mk("Modal Cases - Eigen","Modal Case Definitions - Eigen",
           ["Name","Exclude Group","Mass Source","Stiffness Type","Max Modes","Min Modes",
            "Freq Shift","Cutoff Freq","Convergence Tol","Auto Shift?","Design Type","GUID","Notes"],
           [None,None,None,None,None,None,"cyc/sec","cyc/sec",None,None,None,None,None],
           [["Modal",None,"MsSrc1","P-Delta",12,1,0,0,0,"Yes","Program Determined",str(uuid.uuid4()),None]])

        mk("Mass Source Definition","Mass Source Definition",
           ["Name","Is Default","Include Lateral Mass?","Include Vertical Mass?","Lump Mass?",
            "Source Self Mass?","Source Added Mass?","Source Load Patterns?","Move Mass Centroid?","GUID"],
           [None]*10,
           [["MsSrc1","Yes","Yes","No","Yes","Yes","Yes","No","No",str(uuid.uuid4())]])

        mk("Load Combination Definitions","Load Combination Definitions",
           ["Name","Type","Is Auto","Load Name","SF","GUID","Notes"],[None]*7,
           [["SLS: D","Linear Add","No","Dead",1,str(uuid.uuid4()),None],
            ["SLS: D",None,None,"Super Dead Load",1,None,None],
            ["SLS: D+L","Linear Add","No","Dead",1,str(uuid.uuid4()),None],
            ["SLS: D+L",None,None,"Super Dead Load",1,None,None],
            ["SLS: D+L",None,None,"Live",1,None,None],
            ["ULS: 1.2D+1.6L","Linear Add","No","Dead",1.2,str(uuid.uuid4()),None],
            ["ULS: 1.2D+1.6L",None,None,"Live",1.6,None,None],
            ["ULS: 1.2D+1.6L",None,None,"Super Dead Load",1.2,None,None],
            ["ULS: 1.2D+L+W","Linear Add","No","Dead",1.2,str(uuid.uuid4()),None],
            ["ULS: 1.2D+L+W",None,None,"Super Dead Load",1.2,None,None],
            ["ULS: 1.2D+L+W",None,None,"Live",1,None,None],
            ["ULS: 1.2D+L+W",None,None,"Wind",1,None,None],
            ["ULS: D+0.9W","Linear Add","No","Dead",1,str(uuid.uuid4()),None],
            ["ULS: D+0.9W",None,None,"Super Dead Load",1,None,None],
            ["ULS: D+0.9W",None,None,"Wind",0.9,None,None],
            ["ULS: D+L+W","Linear Add","No","Dead",1,str(uuid.uuid4()),None],
            ["ULS: D+L+W",None,None,"Super Dead Load",1,None,None],
            ["ULS: D+L+W",None,None,"Live",1,None,None],
            ["ULS: D+L+W",None,None,"Wind",1,None,None]])

        mk("Wall Property Def - Specified","Wall Property Definitions - Specified",
           ["Name","Modeling Type","Material","Wall Thickness","Include Auto Rigid Zone?",
            "Notional Size Type","Notional Auto Factor","f11 Modifier","f22 Modifier","f12 Modifier",
            "m11 Modifier","m22 Modifier","m12 Modifier","v13 Modifier","v23 Modifier",
            "Mass Modifier","Weight Modifier","Color","GUID","Notes"],
           [None,None,None,"mm",None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None],
           [["W300","Shell-Thin","4000Psi",300,"No","Auto",1,1,1,1,1,1,1,1,1,1,1,"Red",
             "c0b2c725-19a3-4590-a8e3-49f360b0073b",None]])

        mk("Slab Property Definitions","Slab Property Definitions",
           ["Name","Modeling Type","Property Type","Material","Slab Thickness",
            "One Way Load Distribution?","Notional Size Type","Notional Auto Factor",
            "f11 Modifier","f22 Modifier","f12 Modifier","m11 Modifier","m22 Modifier",
            "m12 Modifier","v13 Modifier","v23 Modifier","Mass Modifier","Weight Modifier",
            "Color","GUID","Notes"],
           [None]*21,
           [["Slab1","Shell-Thin","Slab","4000Psi",200,None,"Auto",1,1,1,1,1,1,1,1,1,1,1,
             "Yellow",str(uuid.uuid4()),None]])

        mk("Shell Uniform Load Sets","Shell Uniform Load Sets",
           ["Name","Load Pattern","Load Value","GUID"],[None,None,"kN/m²",None],
           [["Slab Loading","Super Dead Load",0.72,str(uuid.uuid4())],
            ["Slab Loading","Live",1.92,None]])

        # Reorder sheets
        order = ["Program Control","Grid Definitions - General","Grid Definitions - Grid Lines",
                 "Story Definitions","Tower and Base Story Definition","Load Pattern Definitions",
                 "Load Cases - Summary","Load Cases - Linear Static","Modal Cases - Eigen",
                 "Mass Source Definition","Load Combination Definitions",
                 "Wall Property Def - Specified","Slab Property Definitions","Shell Uniform Load Sets",
                 "Point Object Connectivity","Joint Assigns - Restraints",
                 "Wall Object Connectivity","Objects and Elements - Areas","Area Assigns - Sect Prop"]
        existing = wb.sheetnames
        for i, sn in enumerate([s for s in order if s in existing] +
                                [s for s in existing if s not in order]):
            wb.move_sheet(sn, offset=i - wb.sheetnames.index(sn))

        wb.save(xlsx_path)
        print(f"  ETABS xlsx saved: {xlsx_path}")

    except ImportError:
        logger.warning("openpyxl not installed. Run: pip install openpyxl")
    except Exception as e:
        import traceback
        logger.warning("ETABS xlsx export failed: %s", e)
        traceback.print_exc()


# -----------------------------------------------------------------------
# Export best plan to ETABS-importable xlsx
# (uses same logic as json_to_etabs_xlsx.py)
# -----------------------------------------------------------------------

def _export_etabs_xlsx(json_path: Path, xlsx_path: Path, template_path=None) -> None:
    """Convert before-format JSON to ETABS-importable xlsx."""
    try:
        import json, ast, math, uuid as _uuid
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Font

        # ---- Load JSON ----
        with open(json_path) as f:
            data = json.load(f)
        fd      = data["File_details"][0]
        stories = sorted(fd["Story_detail"], key=lambda s: s["Elevation"])
        floor_h = 4.0

        # ---- Build geometry ----
        joints        = {}
        jc            = [1]
        wall_list     = []
        story_elevs   = []

        def get_jid(coord):
            key = (round(coord[0],3), round(coord[1],3), round(coord[2],3))
            if key not in joints: joints[key] = jc[0]; jc[0] += 1
            return joints[key]

        def parse_jt(s):
            v = ast.literal_eval(str(s)); return (float(v[0]),float(v[1]),float(v[2]))

        for story in stories:
            sname = story["Story"]; elev = story["Elevation"]
            story_elevs.append((sname, elev))
            for w in story.get("Wall_Details", []):
                j1=parse_jt(w["Elm Jt1"]); j2=parse_jt(w["Elm Jt2"])
                j3=parse_jt(w["Elm Jt3"]); j4=parse_jt(w["Elm Jt4"])
                wall_list.append({
                    "story": sname, "elem_name": w["Element Name"],
                    "obj_label": w["Object Label"], "obj_name": w["Object Name"],
                    "jt1": get_jid(j1), "jt2": get_jid(j2),
                    "jt3": get_jid(j3), "jt4": get_jid(j4),
                })

        id2coord = {v: k for k, v in joints.items()}

        def story_for_z(z):
            for sn, el in story_elevs:
                if abs(z - el) < 0.01: return sn
            return "Base"

        # Infer layout
        lx = max(1, int(round(max(x for (x,y,z) in joints) / 6)))
        ly = max(1, int(round(max(y for (x,y,z) in joints) / 6)))

        # ---- Workbook: load template or create fresh ----
        if template_path and Path(template_path).exists():
            wb = load_workbook(str(template_path))
            # Remove geometry sheets that will be regenerated
            for sn in ["Grid Definitions - Grid Lines", "Story Definitions",
                       "Point Object Connectivity", "Joint Assigns - Restraints",
                       "Wall Object Connectivity", "Objects and Elements - Areas",
                       "Area Assigns - Sect Prop"]:
                if sn in wb.sheetnames:
                    del wb[sn]
        else:
            wb = Workbook()
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]

        # ---- Helper: create sheet in ETABS TABLE format ----
        def mk(name, title, headers, units, rows):
            if name in wb.sheetnames:
                del wb[name]
            ws = wb.create_sheet(name)
            ws.append([f"TABLE:  {title}"])
            ws.append(headers)
            ws.append(units)
            for r in rows: ws.append(r)
            ws["A1"].font = Font(bold=True)

        def _sheet(name, title, headers, units, rows):
            """Only create if not already in workbook (from template)."""
            if name not in wb.sheetnames:
                mk(name, title, headers, units, rows)

        # ---- 1. Grid Definitions - General ----
        _sheet("Grid Definitions - General", "Grid Definitions - General",
               ["Name","Type","Ux","Uy","Rz","Story Range","Bubble Size","Color","GUID"],
               [None,None,"m","m","deg",None,"mm",None,None],
               [["G1","Cartesian",0,0,0,"Default",1250,"Gray6",
                 "198239d6-89fe-48ba-8237-939aad37fdc2"]])

        # ---- 2. Grid Definitions - Grid Lines ----
        xl = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
        grid_rows  = [["G1","X (Cartesian)",xl[i],i*6,"End","Yes"] for i in range(lx+1)]
        grid_rows += [["G1","Y (Cartesian)",str(j+1),j*6,"Start","Yes"] for j in range(ly+1)]
        mk("Grid Definitions - Grid Lines", "Grid Definitions - Grid Lines",
           ["Name","Grid Line Type","ID","Ordinate","Bubble Location","Visible"],
           [None,None,None,"m",None,None], grid_rows)

        # ---- 3. Story Definitions ----
        colors = ["Yellow","Gray8Dark","Red","Blue","Green","Orange"]
        rev    = list(reversed(story_elevs))
        master = rev[0][0]
        mk("Story Definitions", "Story Definitions",
           ["Name","Height","Master Story","Similar To","Splice Story","Color","GUID"],
           [None,"m",None,None,None,None,None],
           [[sn, floor_h, "Yes" if i==0 else "No", master if i>0 else None,
             "No", colors[i%len(colors)], str(_uuid.uuid4())]
            for i,(sn,_) in enumerate(rev)])

        # ---- 4. Point Object Connectivity ----
        def fmt(v): return int(v) if v == int(v) else v
        mk("Point Object Connectivity", "Point Object Connectivity",
           ["UniqueName","Is Auto Point","Story","PointBay","IsSpecial","X","Y","Z","GUID"],
           [None,None,None,None,None,"m","m","m",None],
           [[jid,"No",story_for_z(id2coord[jid][2]),jid,"No",
             fmt(id2coord[jid][0]),fmt(id2coord[jid][1]),fmt(id2coord[jid][2]),None]
            for jid in range(1, len(joints)+1)])

        # ---- 5. Joint Assigns - Restraints (Base only) ----
        rc = 1; rest_rows = []
        for jid in range(1, len(joints)+1):
            if id2coord[jid][2] < 0.01:
                rest_rows.append(["Base",jid,rc,"Yes","Yes","Yes","Yes","Yes","Yes"]); rc+=1
        mk("Joint Assigns - Restraints", "Joint Assignments - Restraints",
           ["Story","Label","UniqueName","UX","UY","UZ","RX","RY","RZ"],
           [None]*9, rest_rows)

        # ---- 6. Wall Object Connectivity ----
        mk("Wall Object Connectivity", "Wall Object Connectivity",
           ["UniqueName","Story","WallBay","UniquePt1","UniquePt2",
            "UniquePt3","UniquePt4","Perimeter","Area","GUID"],
           [None,None,None,None,None,None,None,"m","m²",None],
           [[w["elem_name"],w["story"],"Wall",
             w["jt1"],w["jt2"],w["jt3"],w["jt4"],None,None,None]
            for w in wall_list])

        # ---- 7. Objects and Elements - Areas ----
        mk("Objects and Elements - Areas", "Objects and Elements - Areas",
           ["Story","Element Name","Object Type","Object Label","Object Name",
            "Elm Jt1","Elm Jt2","Elm Jt3","Elm Jt4"],
           [None]*9,
           [[w["story"],w["elem_name"],"Wall",w["obj_label"],w["obj_name"],
             w["jt1"],w["jt2"],w["jt3"],w["jt4"]] for w in wall_list])

        # ---- 8. Area Assigns - Sect Prop ----
        mk("Area Assigns - Sect Prop", "Area Assignments - Section Properties",
           ["Story","Label","UniqueName","Section Property","Property Type"],
           [None]*5,
           [[w["story"],"Wall",w["elem_name"],"W300","Wall"] for w in wall_list])

        # ---- Fixed sheets (only if not from template) ----
        _sheet("Program Control", "Program Control",
               ["ProgramName","Version","ProgLevel","LicenseNum","CurrUnits",
                "StlFrmCode","CompBmCode","CompColCode","StlJstCode",
                "ConcFrmCode","ConcSlbCode","ShrWallCode"],
               [None]*12,
               [["ETABS","21.2.0","Ultimate C","3010-*1BTXCCSZTGUALPH","kip, in, F",
                 "AISC 360-16","AISC 360-16","AISC 360-16","SJI-2010",
                 "ACI 318-19","ACI 318-19","ACI 318-19"]])
        _sheet("Tower and Base Story Definition","Tower and Base Story Definitions",
               ["Tower","Color","GUID","Notes","BSName","BSElev","BSColor"],
               [None,None,None,None,None,"m",None],
               [["T1","Gray8Dark",str(_uuid.uuid4()),None,"Base",0,"Gray8Dark"]])
        _sheet("Load Pattern Definitions","Load Pattern Definitions",
               ["Name","Is Auto Load","Type","Self Weight Multiplier","Auto Load","GUID"],
               [None]*6,
               [["~ChineseX","Yes","Other",0,None,str(_uuid.uuid4())],
                ["~ChineseY","Yes","Other",0,None,str(_uuid.uuid4())],
                ["~LLRF","Yes","Other",0,None,str(_uuid.uuid4())],
                ["Dead","No","Dead",1,None,str(_uuid.uuid4())],
                ["Live","No","Live",0,None,str(_uuid.uuid4())],
                ["Super Dead Load","No","Super Dead",0,None,
                 "1ac47084-22cf-4ef7-9577-8308fe7bedcc"],
                ["Wind","No","Wind",0,None,"65671afc-9585-4847-841d-bf6144d0a108"]])
        _sheet("Load Cases - Summary","Load Case Definitions - Summary",
               ["Name","Type","GUID"],[None]*3,
               [["Dead","Linear Static",str(_uuid.uuid4())],
                ["1.2D+L+W","Linear Static","abdb0c4b-a9c4-4dc1-b422-76ab982aa8da"],
                ["Super Dead Load","Linear Static","797f76d0-e4d9-4e5a-b89c-0525aca1e2f7"],
                ["Wind","Linear Static","e962596b-8d4c-4da9-8dc2-47d352ebb8b6"],
                ["Live","Linear Static",str(_uuid.uuid4())],
                ["Modal","Modal - Eigen",str(_uuid.uuid4())]])
        _sheet("Load Cases - Linear Static","Load Case Definitions - Linear Static",
               ["Name","Exclude Group","Mass Source","Stiffness Type","Load Type",
                "Load Name","Load SF","Design Type","GUID","Notes"],[None]*10,
               [["1.2D+L+W",None,"MsSrc1","P-Delta","Load","Dead",1.2,
                 "Program Determined","abdb0c4b-a9c4-4dc1-b422-76ab982aa8da",None],
                ["1.2D+L+W",None,None,None,"Load","Super Dead Load",1.2,None,None,None],
                ["1.2D+L+W",None,None,None,"Load","Live",1,None,None,None],
                ["1.2D+L+W",None,None,None,"Load","Wind",1,None,None,None],
                ["Dead",None,"MsSrc1","P-Delta","Load","Dead",1,
                 "Program Determined",str(_uuid.uuid4()),None],
                ["Live",None,"MsSrc1","P-Delta","Load","Live",1,
                 "Program Determined",str(_uuid.uuid4()),None],
                ["Super Dead Load",None,"MsSrc1","P-Delta","Load","Super Dead Load",1,
                 "Program Determined","797f76d0-e4d9-4e5a-b89c-0525aca1e2f7",None],
                ["Wind",None,"MsSrc1","P-Delta","Load","Wind",1,
                 "Program Determined","e962596b-8d4c-4da9-8dc2-47d352ebb8b6",None]])
        _sheet("Modal Cases - Eigen","Modal Case Definitions - Eigen",
               ["Name","Exclude Group","Mass Source","Stiffness Type","Max Modes","Min Modes",
                "Freq Shift","Cutoff Freq","Convergence Tol","Auto Shift?","Design Type",
                "GUID","Notes"],
               [None,None,None,None,None,None,"cyc/sec","cyc/sec",None,None,None,None,None],
               [["Modal",None,"MsSrc1","P-Delta",12,1,0,0,0,"Yes",
                 "Program Determined",str(_uuid.uuid4()),None]])
        _sheet("Mass Source Definition","Mass Source Definition",
               ["Name","Is Default","Include Lateral Mass?","Include Vertical Mass?",
                "Lump Mass?","Source Self Mass?","Source Added Mass?",
                "Source Load Patterns?","Move Mass Centroid?","GUID"],
               [None]*10,
               [["MsSrc1","Yes","Yes","No","Yes","Yes","Yes","No","No",str(_uuid.uuid4())]])
        _sheet("Load Combination Definitions","Load Combination Definitions",
               ["Name","Type","Is Auto","Load Name","SF","GUID","Notes"],[None]*7,
               [["SLS: D","Linear Add","No","Dead",1,str(_uuid.uuid4()),None],
                ["SLS: D",None,None,"Super Dead Load",1,None,None],
                ["SLS: D+L","Linear Add","No","Dead",1,str(_uuid.uuid4()),None],
                ["SLS: D+L",None,None,"Super Dead Load",1,None,None],
                ["SLS: D+L",None,None,"Live",1,None,None],
                ["ULS: 1.2D+1.6L","Linear Add","No","Dead",1.2,str(_uuid.uuid4()),None],
                ["ULS: 1.2D+1.6L",None,None,"Live",1.6,None,None],
                ["ULS: 1.2D+1.6L",None,None,"Super Dead Load",1.2,None,None],
                ["ULS: 1.2D+L+W","Linear Add","No","Dead",1.2,str(_uuid.uuid4()),None],
                ["ULS: 1.2D+L+W",None,None,"Super Dead Load",1.2,None,None],
                ["ULS: 1.2D+L+W",None,None,"Live",1,None,None],
                ["ULS: 1.2D+L+W",None,None,"Wind",1,None,None],
                ["ULS: D+0.9W","Linear Add","No","Dead",1,str(_uuid.uuid4()),None],
                ["ULS: D+0.9W",None,None,"Super Dead Load",1,None,None],
                ["ULS: D+0.9W",None,None,"Wind",0.9,None,None],
                ["ULS: D+L+W","Linear Add","No","Dead",1,str(_uuid.uuid4()),None],
                ["ULS: D+L+W",None,None,"Super Dead Load",1,None,None],
                ["ULS: D+L+W",None,None,"Live",1,None,None],
                ["ULS: D+L+W",None,None,"Wind",1,None,None]])
        _sheet("Wall Property Def - Specified","Wall Property Definitions - Specified",
               ["Name","Modeling Type","Material","Wall Thickness",
                "Include Auto Rigid Zone?","Notional Size Type","Notional Auto Factor",
                "f11 Modifier","f22 Modifier","f12 Modifier","m11 Modifier",
                "m22 Modifier","m12 Modifier","v13 Modifier","v23 Modifier",
                "Mass Modifier","Weight Modifier","Color","GUID","Notes"],
               [None,None,None,"mm"]+[None]*16,
               [["W300","Shell-Thin","4000Psi",300,"No","Auto",1,1,1,1,1,1,1,1,1,1,1,
                 "Red","c0b2c725-19a3-4590-a8e3-49f360b0073b",None],
                ["Wall1","Shell-Thin","4000Psi",250,"No","Auto",1,1,1,1,1,1,1,1,1,1,1,
                 "Blue",str(_uuid.uuid4()),None]])
        _sheet("Slab Property Definitions","Slab Property Definitions",
               ["Name","Modeling Type","Property Type","Material","Slab Thickness",
                "One Way Load Distribution?","Notional Size Type","Notional Auto Factor",
                "f11 Modifier","f22 Modifier","f12 Modifier","m11 Modifier","m22 Modifier",
                "m12 Modifier","v13 Modifier","v23 Modifier","Mass Modifier",
                "Weight Modifier","Color","GUID","Notes"],
               [None,None,None,None,"mm"]+[None]*16,
               [["Plank1","Membrane","Slab","4000Psi",200,"Yes","Auto",1,1,1,1,1,1,1,1,1,1,1,
                 "Green",str(_uuid.uuid4()),None],
                ["S200","Shell-Thin","Slab","4000Psi",200,None,"Auto",1,1,1,1,1,1,1,1,1,1,1,
                 "Yellow",str(_uuid.uuid4()),None],
                ["Slab1","Shell-Thin","Slab","4000Psi",200,None,"Auto",1,1,1,1,1,1,1,1,1,1,1,
                 "Yellow",str(_uuid.uuid4()),None]])
        _sheet("Shell Uniform Load Sets","Shell Uniform Load Sets",
               ["Name","Load Pattern","Load Value","GUID"],[None,None,"kN/m²",None],
               [["Slab Loading","Super Dead Load",0.72,str(_uuid.uuid4())],
                ["Slab Loading","Live",1.92,None]])

        # ---- Reorder sheets ----
        desired = [
            "Program Control","Grid Definitions - General","Grid Definitions - Grid Lines",
            "Story Definitions","Tower and Base Story Definition","Load Pattern Definitions",
            "Load Cases - Summary","Load Cases - Linear Static","Modal Cases - Eigen",
            "Mass Source Definition","Load Combination Definitions",
            "Wall Property Def - Specified","Slab Property Definitions",
            "Shell Uniform Load Sets","Point Object Connectivity",
            "Joint Assigns - Restraints","Wall Object Connectivity",
            "Objects and Elements - Areas","Area Assigns - Sect Prop",
        ]
        existing = wb.sheetnames
        ordered  = [s for s in desired if s in existing]
        extra    = [s for s in existing if s not in ordered]
        for i, sn in enumerate(ordered + extra):
            wb.move_sheet(sn, offset=i - wb.sheetnames.index(sn))

        wb.save(str(xlsx_path))
        logger.info("ETABS xlsx saved: %s  (%d sheets)", xlsx_path.name, len(wb.sheetnames))

    except ImportError:
        logger.warning("openpyxl not installed. Run: pip install openpyxl")
    except Exception as e:
        import traceback
        logger.warning("ETABS xlsx export failed: %s", e)
        traceback.print_exc()


# -----------------------------------------------------------------------
# Export best plan to before-format JSON
# -----------------------------------------------------------------------

def _export_before_json(walls, lx, ly, num_floors, shear_ratio, out_path, floor_height=4.0):
    """
    Export best plan walls to JSON matching the 'before' dataset format.
    X-Dir and Y-Dir set to 0.0 (filled by ETABS after analysis).
    Wall corners: Jt1=bottom-start, Jt2=bottom-end, Jt3=top-end, Jt4=top-start.
    """
    import json, math

    story_details = []
    for floor_idx in range(num_floors):
        elevation    = (floor_idx + 1) * floor_height
        z_bot        = elevation - floor_height
        z_top        = elevation
        wall_details = []

        for w_idx, w in enumerate(walls):
            cx, cy = w["cx"], w["cy"]
            ang    = w["angle"]
            half   = w.get("length", 2.0) / 2.0
            dx     = round(half * math.cos(math.radians(ang)), 3)
            dy     = round(half * math.sin(math.radians(ang)), 3)
            x1, y1 = round(cx - dx, 3), round(cy - dy, 3)
            x2, y2 = round(cx + dx, 3), round(cy + dy, 3)

            wall_details.append({
                "Element Name": w_idx + 1,
                "Object Type":  "Wall",
                "Object Label": f"W{w_idx + 1}",
                "Object Name":  w_idx + 1,
                "Elm Jt1": str([int(x1) if x1 == int(x1) else x1,
                          int(y1) if y1 == int(y1) else y1,
                          int(z_bot) if z_bot == int(z_bot) else z_bot]),
                "Elm Jt2": str([int(x2) if x2 == int(x2) else x2,
                          int(y2) if y2 == int(y2) else y2,
                          int(z_bot) if z_bot == int(z_bot) else z_bot]),
                "Elm Jt3": str([int(x2) if x2 == int(x2) else x2,
                          int(y2) if y2 == int(y2) else y2,
                          int(z_top) if z_top == int(z_top) else z_top]),
                "Elm Jt4": str([int(x1) if x1 == int(x1) else x1,
                          int(y1) if y1 == int(y1) else y1,
                          int(z_top) if z_top == int(z_top) else z_top]),
            })

        story_details.append({
            "Story":        f"Story{floor_idx + 1}",
            "Elevation":    round(elevation, 3),
            "Location":     "Top",
            "X-Dir":        0.0,
            "Y-Dir":        0.0,
            "Wall_Details": wall_details,
        })

    # Reverse story order: top floor first (matching original before format)
    story_details = list(reversed(story_details))

    data = {
        "ProjectName": f"Project_{lx}x{ly}_{num_floors}",
        "File_details": [{
            "Name of file":     f"generated_{lx}x{ly}_f{num_floors}.json",
            "Shear_wall_ratio": float(shear_ratio),
            "Story_detail":     story_details,
        }],
    }
    with open(out_path, "w") as f:
        json.dump(data, f, indent=2)
    print(f"  Before-format JSON saved: {out_path}")


# -----------------------------------------------------------------------
# Strict canonical drift prediction & ranking
#
# Workflow (no in-process reimplementation; uses doctor's exact scripts):
#   Step 1: Save each candidate's walls as a before-format JSON file
#   Step 2: Run O1graph_converter_v3.build_enhanced_graph for each JSON,
#           save resulting Data object to .pt
#   Step 3: Run O3inference_predict_only.predict_for_pipeline on graphs
#           (no GT required at this phase; ETABS analysis happens later)
#           This script imports model arch + helpers from O3inference_v9.py
#   Step 4: Parse per-candidate xlsx output to extract drift predictions
#   Step 5: Rank candidates by mean total drift (lower = better)
#
# All intermediate artifacts are stored under per_cand_dir for inspection
# and reproducibility.
# -----------------------------------------------------------------------

def _rerank_canonical_strict(all_candidates, lx, ly, num_floors, shear_ratio,
                             drift_ckpt, graph_converter_path,
                             inference_script_path, per_cand_dir,
                             timing_dict=None, device="cpu"):
    """
    Re-rank candidates using the doctor's exact canonical pipeline.

    Args:
        all_candidates:        list of candidate dicts; each must contain "walls"
        lx, ly:                plan dimensions (number of bays in X and Y)
        num_floors:            number of stories
        shear_ratio:           target shear wall ratio (used in JSON metadata)
        drift_ckpt:            path to v9 drift surrogate checkpoint
        graph_converter_path:  path to O1graph_converter_v3.py
        inference_script_path: path to O3inference_predict_only.py
                               (which itself imports model arch + helpers
                               from O3inference_v9.py in the same folder)
        per_cand_dir:          workspace directory for intermediate files

    Returns:
        list of candidates sorted by ascending mean total drift,
        each augmented with "drift_pred_norm", "x_drift_mean",
        "y_drift_mean", "x_drift_norm_mean", "y_drift_norm_mean".
    """
    import importlib.util
    from pathlib import Path
    import torch
    import pandas as pd
    import numpy as np
    import time

    _meas = timing_dict is not None
    _t_predict = 0.0   # Step 2 + Step 3
    _t_rerank  = 0.0   # Step 1 + Step 4 + Step 5

    per_cand_dir = Path(per_cand_dir)
    json_dir      = per_cand_dir / "jsons"
    graph_dir     = per_cand_dir / "graphs"
    inference_dir = per_cand_dir / "inference"
    json_dir.mkdir(parents=True, exist_ok=True)
    graph_dir.mkdir(parents=True, exist_ok=True)
    inference_dir.mkdir(parents=True, exist_ok=True)

    n = len(all_candidates)
    logger.info("=" * 60)
    logger.info("Strict canonical drift workflow (%d candidates)", n)
    logger.info("=" * 60)

    # ------------------------------------------------------------------
    # Step 1: Save each candidate's walls as a before-format JSON
    # ------------------------------------------------------------------
    logger.info("Step 1: Saving candidate JSONs to %s", json_dir)
    _ts = time.perf_counter() if _meas else 0.0
    for idx, cand in enumerate(all_candidates):
        json_path = json_dir / f"before_candidate_{idx+1:03d}.json"
        _export_before_json(
            walls=cand.get("walls", []),
            lx=lx, ly=ly, num_floors=num_floors,
            shear_ratio=shear_ratio,
            out_path=json_path,
        )
    if _meas:
        _t_rerank += time.perf_counter() - _ts

    # ------------------------------------------------------------------
    # Step 2: Build supergraphs via canonical O1graph_converter_v3.py
    # ------------------------------------------------------------------
    logger.info("Step 2: Building supergraphs via O1graph_converter_v3.py")
    _ts = time.perf_counter() if _meas else 0.0
    spec = importlib.util.spec_from_file_location(
        "o1_graph_converter", str(graph_converter_path)
    )
    gc = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(gc)

    for idx in range(n):
        json_path  = json_dir  / f"before_candidate_{idx+1:03d}.json"
        graph_path = graph_dir / f"supergraph_candidate_{idx+1:03d}.pt"
        graph, _, _ = gc.build_enhanced_graph(str(json_path))
        torch.save(graph, graph_path)
        logger.info("  Built %s (%d nodes)", graph_path.name, graph.x.size(0))
    if _meas:
        _t_predict += time.perf_counter() - _ts

    # ------------------------------------------------------------------
    # Step 3: Run predict-only inference (NO GT required)
    # Uses O3inference_predict_only.py, which imports the canonical model
    # architecture and helpers from O3inference_v9.py but skips all
    # GT-based evaluation (MAPE/R^2/plots). This is the appropriate
    # entrypoint for Phase-1 inference: at this point GT does not yet
    # exist (ETABS analysis happens later, in Phase 2).
    # ------------------------------------------------------------------
    logger.info("Step 3: Running %s.predict_for_pipeline",
                Path(inference_script_path).stem)
    predict_script = Path(inference_script_path)
    if not predict_script.exists():
        raise FileNotFoundError(
            f"Inference script not found: {predict_script}\n"
            f"Provide a valid path via --inference_script. Note: that script "
            f"in turn requires O3inference_v9.py in the same folder."
        )

    spec_p = importlib.util.spec_from_file_location(
        "o3inference_predict_only", str(predict_script)
    )
    pmod = importlib.util.module_from_spec(spec_p)
    spec_p.loader.exec_module(pmod)

    if _meas and device == "cuda":
        torch.cuda.synchronize()
    _ts = time.perf_counter() if _meas else 0.0
    run_folder = pmod.predict_for_pipeline(
        model_path=str(drift_ckpt),
        graph_folder=str(graph_dir),
        output_folder=str(inference_dir),
    )
    if _meas:
        if device == "cuda":
            torch.cuda.synchronize()
        _t_predict += time.perf_counter() - _ts

    # ------------------------------------------------------------------
    # Step 4: Parse per-candidate xlsx outputs to extract predictions
    # ------------------------------------------------------------------
    logger.info("Step 4: Parsing predictions from %s", run_folder.name)
    _ts = time.perf_counter() if _meas else 0.0
    for idx, cand in enumerate(all_candidates):
        xlsx_path = run_folder / f"supergraph_candidate_{idx+1:03d}.xlsx"
        if not xlsx_path.exists():
            logger.warning("  Missing xlsx for candidate %d: %s",
                           idx + 1, xlsx_path.name)
            cand["drift_pred_norm"]   = None
            cand["x_drift_norm_mean"] = float("inf")
            cand["y_drift_norm_mean"] = float("inf")
            cand["x_drift_mean"]      = float("inf")
            cand["y_drift_mean"]      = float("inf")
            continue

        df = pd.read_excel(xlsx_path)
        # Filter out Base floor (no SUMMARY in predict-only output)
        df_floors = df[df["Floor"].astype(str) != "Base"]
        # Sort ascending by Elevation (Floor 1 to N)
        df_floors = df_floors.sort_values("Elevation")

        x_pred = df_floors["X-Dir_Pred"].values.astype(float)
        y_pred = df_floors["Y-Dir_Pred"].values.astype(float)

        drift_pred_norm = np.stack([x_pred, y_pred], axis=1)  # [F, 2]
        cand["drift_pred_norm"]   = drift_pred_norm
        cand["x_drift_norm_mean"] = float(x_pred.mean())
        cand["y_drift_norm_mean"] = float(y_pred.mean())
        cand["x_drift_mean"]      = cand["x_drift_norm_mean"]
        cand["y_drift_mean"]      = cand["y_drift_norm_mean"]

        logger.info("  Candidate %d: X_norm_mean=%.2f  Y_norm_mean=%.2f",
                    idx + 1, cand["x_drift_norm_mean"], cand["y_drift_norm_mean"])

    # ------------------------------------------------------------------
    # Step 5: Rank by total mean drift (lower = better)
    # ------------------------------------------------------------------
    logger.info("Step 5: Ranking by total mean drift (lower = better)")
    valid   = [c for c in all_candidates if c.get("drift_pred_norm") is not None]
    invalid = [c for c in all_candidates if c.get("drift_pred_norm") is None]

    for c in valid:
        total = c["x_drift_mean"] + c["y_drift_mean"]
        c["final_score"] = -total  # higher score = lower drift = better

    valid.sort(key=lambda c: c["final_score"], reverse=True)

    # Update rank
    ranked = valid + invalid
    for i, c in enumerate(ranked):
        c["rank"] = i + 1

    if _meas:
        _t_rerank += time.perf_counter() - _ts   # Step 4 + Step 5
        timing_dict["predict"] = _t_predict
        timing_dict["rerank"]  = _t_rerank

    return ranked



# -----------------------------------------------------------------------
# Visualization
# -----------------------------------------------------------------------

def _save_candidate_png(cand, rank, is_best, lx, ly, num_floors, shear_ratio, out_dir):
    """Render a clean 2D floor-plan (top view) of one candidate layout.

    Single-panel figure intended for side-by-side comparison of candidates
    (Fig. gen_candidates, Section 7.3). Walls are coloured by orientation:
    horizontal (along X) in orange, vertical (along Y) in blue, matching the
    interpolation figure. Per-candidate metrics (drift, symmetry, ratio) are
    reported in the accompanying LaTeX table, not embedded in the image.
    Fonts follow the manuscript (Computer Modern / Latin Modern, elsarticle 12pt).
    """
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        import numpy as np

        # --- Font: match LaTeX (CM/LM), no usetex ---
        plt.rcParams.update({
            "font.family":      "serif",
            "font.serif":       ["CMU Serif", "Latin Modern Roman",
                                 "Computer Modern Roman", "DejaVu Serif"],
            "mathtext.fontset": "cm",
            "axes.unicode_minus": False,
        })

        bx, by  = lx * 6.0, ly * 6.0
        x_ticks = list(range(0, int(bx) + 1, 6))
        y_ticks = list(range(0, int(by) + 1, 6))
        # Plan palette: H = orange (along X), V = blue (along Y)
        COLOR_H = "#F4A460"
        COLOR_V = "#1F77B4"

        walls   = cand.get("walls", [])
        n_walls = len(walls)
        max_slots = (ly + 1) * lx * 3 + (lx + 1) * ly * 3
        rho     = n_walls / max(max_slots, 1)

        # Fix the canvas WIDTH (so text scales identically when each panel is
        # rendered at the same LaTeX width) and clamp the HEIGHT to a sensible
        # band. The plan keeps true geometric proportions via set_aspect; very
        # tall or very flat plans get white margin rather than distorted text.
        aspect = by / bx
        fig_w  = 4.2
        fig_h  = fig_w * min(max(aspect, 0.45), 1.30) + 0.9
        fig, ax = plt.subplots(figsize=(fig_w, fig_h))

        ax.set_facecolor("white")
        for x in x_ticks:
            ax.axvline(x, color="#cccccc", lw=0.6, zorder=0)
        for y in y_ticks:
            ax.axhline(y, color="#cccccc", lw=0.6, zorder=0)

        # Fixed thin line width for all plans (uniform visual weight,
        # independent of plan size). A single value is used for every wall.
        lw_w = 2.0

        for w in walls:
            ang  = w["angle"]
            is_h = abs(np.cos(np.radians(ang))) > abs(np.sin(np.radians(ang)))
            ax.plot([w["x1"], w["x2"]], [w["y1"], w["y2"]],
                    color=COLOR_H if is_h else COLOR_V,
                    lw=lw_w, solid_capstyle="round", zorder=2)

        pad = max(1.0, 0.06 * max(bx, by))
        ax.set_xlim(-pad, bx + pad)
        ax.set_ylim(-pad, by + pad)
        ax.set_xticks(x_ticks)
        ax.set_yticks(y_ticks)
        ax.set_aspect("equal")               # true geometric proportions
        ax.set_xlabel("$X$ (m)", fontsize=9)
        ax.set_ylabel("$Y$ (m)", fontsize=9)
        ax.tick_params(labelsize=8)

        ax.set_title(f"$\\rho={rho:.2f}$,  $n={n_walls}$",
                     fontsize=10)

        plt.tight_layout()
        fname = f"rank{rank:02d}_BEST.png" if is_best else f"rank{rank:02d}.png"
        plt.savefig(out_dir / fname, dpi=300)
        plt.savefig(out_dir / fname.replace(".png", ".pdf"))
        plt.close()
        print(f"  Saved: {fname} (+ pdf)")

    except Exception as e:
        import traceback
        print(f"PNG failed rank {rank}: {e}")
        traceback.print_exc()

def _save_comparison_table(candidates, lx, ly, out_dir, use_drift=True):
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt

        n         = len(candidates)
        max_slots = (ly+1)*lx*3 + (lx+1)*ly*3

        fig, ax = plt.subplots(figsize=(16, max(3, n*0.7+2)))
        ax.axis("off")

        drift_src = "drift_model_norm" if use_drift else "surrogate"
        col_labels = [
            "Rank", "Plan", "Final Score",
            f"X-Drift\n({drift_src})",
            f"Y-Drift\n({drift_src})",
            "Shear Ratio", "Symmetry", "Construct.", "Status"
        ]

        rows = []
        for i, cand in enumerate(candidates):
            walls     = cand.get("walls", [])
            act_ratio = len(walls) / max(max_slots, 1)
            fs        = cand.get("final_score", cand.get("ranking_score", 0))
            rows.append([
                f"#{i+1}", f"{lx}x{ly}",
                f"{fs:.4f}",
                f"{cand['x_drift_mean']:.2f}",
                f"{cand['y_drift_mean']:.2f}",
                f"{act_ratio:.3f}",
                f"{cand.get('symmetry_pred',0):.3f}",
                f"{cand.get('constructability',0):.3f}",
                "BEST" if i == 0 else "OK",
            ])

        table = ax.table(cellText=rows, colLabels=col_labels, loc="center", cellLoc="center")
        table.auto_set_font_size(False)
        table.set_fontsize(10)
        table.scale(1, 1.8)

        for j in range(len(col_labels)):
            table[0,j].set_facecolor("#2c3e50")
            table[0,j].set_text_props(color="white", fontweight="bold")
        for i in range(n):
            for j in range(len(col_labels)):
                if i == 0:
                    table[i+1,j].set_facecolor("#fef9e7")
                    table[i+1,j].set_text_props(fontweight="bold")
                elif i % 2 == 0:
                    table[i+1,j].set_facecolor("#f8f9fa")

        ax.set_title(
            f"Candidate Comparison - Plan {lx}x{ly}  ({n} candidates)\n"
            f"Ranked by: X-drift + Y-drift (lower = better)  |  Source: GATv2 Drift Model (Normalized)",
            fontsize=11, fontweight="bold", pad=20,
        )
        plt.tight_layout()
        plt.savefig(out_dir / "comparison_table.png", dpi=150, bbox_inches="tight")
        plt.close()
        print(f"  Saved: comparison_table.png")

    except Exception as e:
        print(f"Comparison table failed: {e}")


# -----------------------------------------------------------------------
# Argument parser
# -----------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser("Discrete Diffusion Wall Layout Pipeline")
    sub    = parser.add_subparsers(dest="command")

    p_mask = sub.add_parser("build_masks")
    p_mask.add_argument("--data_root",  required=True)
    p_mask.add_argument("--out",        default="./masks.pt")
    p_mask.add_argument("--on_thresh",  type=float, default=0.99)
    p_mask.add_argument("--off_thresh", type=float, default=0.01)

    p_train = sub.add_parser("train")
    p_train.add_argument("--data_root",  required=True)
    p_train.add_argument("--masks",      required=True)
    p_train.add_argument("--epochs",     type=int,   default=100)
    p_train.add_argument("--batch_size", type=int,   default=16)
    p_train.add_argument("--workers",    type=int,   default=4)
    p_train.add_argument("--lr",         type=float, default=1e-4)
    p_train.add_argument("--T",          type=int,   default=200)
    p_train.add_argument("--schedule",   default="cosine")
    p_train.add_argument("--base_ch",    type=int,   default=32)
    p_train.add_argument("--ckpt_dir",   default="outputs/diffusion_checkpoints")
    p_train.add_argument("--resume",     default=None)

    p_gen = sub.add_parser("generate")
    p_gen.add_argument("--masks",            required=True)
    p_gen.add_argument("--ckpt",             required=True)
    p_gen.add_argument("--lx",               type=int,   required=True)
    p_gen.add_argument("--ly",               type=int,   required=True)
    p_gen.add_argument("--floors",           type=int,   default=10)
    p_gen.add_argument("--shear",            type=float, default=0.0)
    p_gen.add_argument("--candidates",       type=int,   default=16)
    p_gen.add_argument("--output_dir",       default="disc_diffusion/outputs/generated")
    p_gen.add_argument("--drift_checkpoint", default=None,
                       help="Path to v9 drift surrogate checkpoint (best_model_v9.pt)")
    p_gen.add_argument("--graph_converter",  default=None,
                       help="Path to O1graph_converter_v3.py (canonical graph builder)")
    p_gen.add_argument("--inference_script", default=None,
                       help="Path to O3inference_predict_only.py (predict-only inference)")
    p_gen.add_argument("--save_before_json", action="store_true",
                       help="Export best plan to before-format JSON for further processing")
    p_gen.add_argument("--save_etabs_xlsx",  action="store_true",
                       help="Export best plan to ETABS-importable xlsx (for ETABS analysis)")
    p_gen.add_argument("--etabs_template",   default=None,
                       help="Template xlsx path (fixed sheets copied from here)")
    p_gen.add_argument("--measure_timing",   action="store_true",
                       help="Measure per-stage timing and append to timing_log.csv")
    p_gen.add_argument("--timing_dir",       default=None,
                       help="Directory to save timing_log.csv (default: output_dir)")

    args = parser.parse_args()

    if args.command == "build_masks":
        cmd_build_masks(args)
    elif args.command == "train":
        cmd_train(args)
    elif args.command == "generate":
        cmd_generate(args)
    else:
        parser.print_help()


if __name__ == "__main__":
    main()